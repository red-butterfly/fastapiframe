import numbers

from openpyxl import Workbook, load_workbook


class ExeclProcesser(object):
    def __init__(self):
        pass

    def load_execl(self, filepath):
        '''
        载入execl
        '''
        self.wb = load_workbook(filepath)

    def load_sheet(self, sheetSign):
        '''
        载入sheet
        sheetSign: 如果为int，则是sheet的编号；如果为str，则是sheet的名称
        '''
        if isinstance(sheetSign, numbers.Integral):
            self.sheet = self.wb.get_sheet_by_name(
                self.wb.get_sheet_names()[sheetSign]
            )
        elif isinstance(sheetSign, str):
            self.sheet = self.wb.get_sheet_by_name(sheetSign)

    def new_execl(self, sheetSign, index=0):
        '''
        新建execl
        '''
        self.wb = Workbook()
        self.wb.create_sheet(sheetSign, index=index)
        self.load_sheet(index)

    def new_sheet(self, sheetSign, index=0):
        '''
        新建工作表
        '''
        self.wb.create_sheet(sheetSign, index=index)

    def save_execl(self, filepath):
        self.wb.save(filepath)

    def get_size(self):
        '''
        返回最大行列 (最大列，最大行)
        '''
        return (self.sheet.max_column, self.sheet.max_row)

    def load_iterdata(self, type='rows', filter=True):
        '''
        按行/列的数据生成器
        'rows' / 'columns'
        '''
        dataiter = self.sheet.rows if type == 'rows' else self.sheet.columns
        max_len = self.sheet.max_row if type == 'rows' else self.sheet.max_column
        haveFilter = False
        for celldata in dataiter:
            if filter and not haveFilter:
                haveFilter = True
                continue
            else:
                yield [cell.value for cell in celldata[0:max_len]]

    def get_cell(self, row, column):
        '''
        获取cell
        '''
        return self.sheet.cell(row=row - 1, column=column - 1)

    def get_linedata(self, index, type='rows'):
        '''
        获取某一行数据
        '''
        dataiter = self.sheet.rows if type == 'rows' else self.sheet.columns
        max_len = self.sheet.max_row if type == 'rows' else self.sheet.max_column
        return [
            cell.value for cell in list(dataiter)[index][0:max_len]
            ]

    def write_line(self, linedata, type='rows'):
        self.sheet.append(linedata)